import os
import sys
import json
import openpyxl.worksheet
import torch
import torch.nn as nn
import torch.utils
import torch.utils.data
from torchvision import transforms, datasets
import torch.optim as optim
from tqdm import tqdm
from model import GoogLeNet
import openpyxl
from openpyxl.chart import LineChart, Reference


# 加载数据集并训练，计算loss和accuracy，保存训练好的网络参数
def main():
    # 如果有NVIDA显卡，转到GPU训练，否则用CPU
    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    print("using {} device.".format(device))

    data_transform = {
        # Compose()：将多个tranforms的操作整合在一起
        # 训练
        "train":
        transforms.Compose([
            # RandomResizedCrop(224)：将给定图像随机裁剪为不同大小和宽高比，然后缩放裁剪所得到的图像为给定大小
            transforms.RandomResizedCrop(224),
            # RandomVerticalFlip()：以0.5的概率竖直翻转给定的PIL图像
            transforms.RandomVerticalFlip(),
            # RandomHorizontalFlip()：以0.5的概率水平翻转给定的PIL图像
            transforms.RandomHorizontalFlip(),
            # ToTensor()：数据转化为Tensor格式
            transforms.ToTensor(),
            # Normalize()：将图像的像素值归一化到[-1,1]之间，使模型容易收敛
            transforms.Normalize((0.5, 0.5, 0.5), (0.5, 0.5, 0.5))
        ]),
        # 验证
        "val":
        transforms.Compose([
            transforms.Resize((224, 224)),
            transforms.ToTensor(),
            transforms.Normalize((0.5, 0.5, 0.5), (0.5, 0.5, 0.5))
        ])
    }

    # abspath()：获取文件当前目录的绝对路径
    # join()：用于拼接文件路径，可以传入多个路径
    # getcwd()：该函数不需要传递参数，获得当前运行的脚本的路径
    data_root = os.path.abspath(os.getcwd())
    # 得到数据集的路径
    image_path = os.path.join(data_root, "fruit")
    # exists():判断括号里的文件是否存在，可以是文件路径
    # 如果image_path不存在，就会抛出AssertionError错误，报错为引号内内容
    assert os.path.exists(image_path), "{} path does not exist.".format(
        image_path)

    # 加载训练数据集集
    # ImageFolder：假设所有的文件按文件夹保存，每个文件夹下存储同一个类别的图片，文件夹名为类名，其构造函数如下
    # ImageFolder(root,
    #             tranform=None,
    #             target_tranform=None,
    #             loader=default_loader)
    # root：在指定路径下寻找图片，tranform：对PILImage进行的转换操作，输入是使用loader读取的图片
    train_dataset = datasets.ImageFolder(root=os.path.join(
        image_path, "train"),
                                         transform=data_transform["train"])
    # 训练集长度
    train_num = len(train_dataset)

    # class_to_idx：获取分类名称对应的索引
    fruit_list = train_dataset.class_to_idx
    # dict()：创建一个新的字典
    # 循环遍历数组索引并交换val和key的值重新赋值给数组，这样模型预测的结果就直接是value类别值
    cla_dict = dict((val, key) for key, val in fruit_list.items())
    # 把字典编码成json格式
    json_str = json.dumps(cla_dict, indent=4)
    # 把字典类别索引写入json文件
    with open('class_indices.json', 'w') as json_file:
        json_file.write(json_str)

    # 一次训练载入32张图像
    batch_size = 32
    # 确定进程数
    # min()：返回给定参数的最小值，参数可以为序列
    # cpu_ count()：返回一个整数值，表示系统中的CPU数亮，如果不确定CPU的数量，则不返回任何内容
    nw = min([os.cpu_count(), batch_size if batch_size > 1 else 0, 8])
    print('Using {} dataloader workers every process'.format(nw))
    # Dataloader：将读取的数据按照batch size大小封装给训练集
    # dataset(Dataset)：输入的数据集
    # batch_size(int, optional)：每个batch加载多少个样本，默认：1
    # shuffle(bool, optional)：设置为True是会在每个epoch重新打乱数据，默认：False
    # num_workers(int, optional)：决定了有几个进程来处理，默认为0，意味着所有数据都会被load进主进程
    train_loader = torch.utils.data.DataLoader(train_dataset,
                                               batch_size=batch_size,
                                               shuffle=True,
                                               num_workers=nw)

    # 加载测试数据集
    validate_dataset = datasets.ImageFolder(root=os.path.join(
        image_path, "val"),
                                            transform=data_transform["val"])

    # 测试集长度
    val_num = len(validate_dataset)
    validate_loader = torch.utils.data.DataLoader(validate_dataset,
                                                  batch_size=batch_size,
                                                  shuffle=False,
                                                  num_workers=nw)

    print("using {} image for training, {} image for validation.".format(
        train_num, val_num))

    # 模型实例化，将模型转到device
    net = GoogLeNet(num_classes=5, aux_logits=True, init_weights=True)
    net.to(device)

    # 定义损失函数（交叉熵损失）
    loss_function = nn.CrossEntropyLoss()

    # 定义adam优化器
    # params(iterable)：要训练的参数，一般传入的是model.parameters()
    # lr(float)：learning rate学习率，也就是步长，默认：1e-3(高了可能过拟合)
    lr = 0.0001
    optimizer = optim.Adam(net.parameters(), lr=lr)

    # 迭代次数（训练次数）
    epochs = 40
    # 用于判断最佳模型
    best_acc = 0.0
    # 最佳模型的保存地址
    save_path = 'googleNet.pth'
    train_steps = len(train_loader)

    # 创建Excel表存储每轮的loss和准确度
    if os.path.exists('训练过程.xlsx'):
        workbook = openpyxl.load_workbook(r'训练过程.xlsx')
    else:
        workbook = openpyxl.Workbook()
    sheet_title = "epoch{} learning rate{}".format(epochs, lr)
    sheet = workbook.create_sheet(sheet_title, 1)
    sheet['A1'] = 'epoch'
    sheet['B1'] = 'loss'
    sheet['C1'] = 'accuracy'

    for epoch in range(epochs):
        # 训练
        net.train()
        running_loss = 0.0
        # tqdm：进度条显示
        train_bar = tqdm(train_loader, file=sys.stdout)
        # train_bar：传入数据（数据包括：训练数据和标签）
        # enumerate()：讲一个可遍历的数据对象（如列表、元组或字符串等）组合为一个索引序列，同时列出数据和数据下标
        # enumerate的返回值有两个：一个是序号，一个是数据（包含训练数据和标签）
        # x：训练数据(input)(tensor类型)，y：标签(labels)(tensor类型)
        for step, data in enumerate(train_bar):
            # 向前传播
            images, labels = data
            # 计算训练值
            logits, aux_logits2, aux_logits1 = net(images.to(device))
            # GoogleNet的网络输出loss有三个部分，分别是主干输出loss、两个辅助分类器输出loss(权重0.3)
            loss0 = loss_function(logits, labels.to(device))
            loss1 = loss_function(aux_logits1, labels.to(device))
            loss2 = loss_function(aux_logits2, labels.to(device))
            loss = loss0 + 0.3 * loss1 + 0.3 * loss2
            # 反向传播
            # 清空过往梯度
            optimizer.zero_grad()
            # 反向传播，计算当前梯度
            loss.backward()
            # 根据梯度更新网络参数
            optimizer.step()
            # item()：得到元素张量的元素值
            running_loss += loss.item()

            # 进度条的前缀
            # .3f：表示浮点的精度为3（小数点后保留3位）
            train_bar.desc = "train epoch[{}/{}] loss:{:.3f}".format(
                epoch + 1, epochs, loss)

        # 测试
        # eval()：如果模型中Batch Normalization和Dropout，则不启用，以防改变权值
        net.eval()
        acc = 0.0
        # 清空历史梯度，与训练的最大区别就是取消了反向传播
        with torch.no_grad():
            val_bar = tqdm(validate_loader, file=sys.stdout)
            for val_data in val_bar:
                val_images, val_labels = val_data
                outputs = net(val_images.to(device))
                # torch.max(input, dim)函数
                # input是具体的tensor，dim是max函数索引的唯独，0是每列的最大值，1是每行的最大值输出
                # 函数会返回两个tensor，第一个tensor是每行的最大值，第二个tensor是每行的最大值的索引
                predict_y = torch.max(outputs, dim=1)[1]
                # 对两个张量tensor进行逐元素的比较，若相同位置的两个元素相同，则返回True；若不同，返回False
                # .sum()对输入的tensor数据的某一维度求和
                acc += torch.eq(predict_y, val_labels.to(device)).sum().item()

        val_accurate = acc / val_num
        print('[epoch %d] train_loss: %.3f val_accuracy: %.3f' %
              (epoch + 1, running_loss / train_steps, val_accurate))

        # 保存最好的模型权重
        if val_accurate > best_acc:
            best_acc = val_accurate
            # torch.save(state, dir)保存模型等的相关参数，dir表示保存的文件路径+保存文件名
            # model.state_dict()：返回的是一个OrderedDict，传出了网络结构的名字和对应的参数
            torch.save(net.state_dict(), save_path)

        # 将每轮的结果保存到Excel里
        sheet.cell(row=epoch + 2, column=1).value = epoch + 1
        sheet.cell(row=epoch + 2, column=2).value = running_loss / train_steps
        sheet.cell(row=epoch + 2, column=3).value = val_accurate

    # category = Reference(sheet,
    #                      min_col=1,
    #                      max_col=1,
    #                      min_row=2,
    #                      max_row=epochs + 1)
    # value1 = Reference(sheet,
    #                    min_col=2,
    #                    max_col=2,
    #                    min_row=2,
    #                    max_row=epochs + 1)
    # value2 = Reference(sheet,
    #                    min_col=3,
    #                    max_col=3,
    #                    min_row=2,
    #                    max_row=epochs + 1)
    # chart = LineChart()
    # chart.add_data(value1)
    # chart.add_data(value2)
    # chart.set_categories(category)
    # chart.title('步长{}'.format(lr))
    workbook.save(r'训练过程.xlsx')
    print('Finished Training')


if __name__ == '__main__':
    main()
